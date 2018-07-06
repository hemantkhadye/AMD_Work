import csv
import requests
from lxml import html
import json
from lxml import html
from openpyxl import load_workbook
import xlsxwriter
from bs4 import BeautifulSoup
input=[]
DataOut=[]

proxy = {'https': 'https://eclerxusa:WonderS1979@Atlanta.wonderproxy.com:80'}
sec=requests.session()
sec.headers = ({'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.186 Safari/537.36',
          'content-type': 'text/html;charset=UTF-8',
           'accept':'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
           'Host': 'www.bestbuy.com'})

#---input file loaded in inputs list
# wb= load_workbook(filename = 'E:\hemant\Hemant Python\SouqConnector\SouqInputs.xlsx')
# ws = wb['Sheet1']
# row = ws.max_row
# col=ws.max_column
# for row in ws.iter_rows(min_row=2, min_col=0, max_row=row, max_col=1):
#     for cell in row:
#         input.append(cell.value)
# wb.close()

#for url in input:
url = "https://www.bestbuy.com/site/searchpage.jsp?cp=2&searchType=search&_dyncharset=UTF-8&ks=960&sc=Global&list=y&usc=All%20Categories&type=page&id=pcat17071&iht=n&seeAll=&browsedCategory=pcmcat309300050015&st=pcmcat309300050015_categoryid%24pcmcat138500050001&qp=features_facet%3DFeatures~2-in-1%20Design%5Econdition_facet%3DCondition~Open-Box%5Econdition_facet%3DCondition~New%22"
CatHome = sec.get(url, proxies=proxy)
soup = BeautifulSoup(CatHome.text,'lxml')
print(soup)
