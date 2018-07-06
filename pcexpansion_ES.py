from tkinter.tix import ResizeHandle
import datetime
import requests
import re
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import xlsxwriter

sess = requests.session()
sess.headers =({
    'Connection': 'keep-alive',
    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36',
    'Upgrade-Insecure-Requests': '1',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
    'Accept-Encoding': 'gzip, deflate, br',
    'Accept-Language': 'en-GB,en;q=0.9,en-US;q=0.8,hi;q=0.7'})
input = []
DataOut = []

def excel_To_List():
    wb= load_workbook(filename='E:\AMD_Work\AMD_Work\Input.xlsx')
    ws = wb['Sheet1']
    row = ws.max_row
    col=ws.max_column
    for row in ws.iter_rows(min_row=2, min_col=0, max_row=row, max_col=1):
        for cell in row:
            input.append(cell.value)
    wb.close()

def Result_Out_Excel(DataOut):
    dt = str(datetime.date.today());
    filename = "PCExpansion_"+country+"_"+dt
    path = "E:\AMD_Work\AMD_Work\\" +filename+".xlsx"
    workbook = xlsxwriter.Workbook(path)
    worksheet = workbook.add_worksheet()
    row = 0
    col = 0

    worksheet.write(row, col, "LOB");
    worksheet.write(row, col+1, "Country");
    worksheet.write(row, col+2, "Site");
    worksheet.write(row, col+3, "Item number");
    worksheet.write(row, col+4, "MPN");
    worksheet.write(row, col+5, "Manufacturer");
    worksheet.write(row, col+6, "ProductName");
    worksheet.write(row, col+7, "ProductURL");
    worksheet.write(row, col+8, "Listprice");
    worksheet.write(row, col+9, "Promoprice");
    worksheet.write(row, col+10, "CurrencyType");
    worksheet.write(row, col+11, "Processor");
    worksheet.write(row, col+12, "RetailerId");
    worksheet.write(row, col+13, "Date");
    row += 1

    for d in DataOut:
        worksheet.write(row, col, LOB);
        worksheet.write(row, col+1, country);
        worksheet.write(row, col+2, site);
        worksheet.write(row, col+3, d[1]);
        worksheet.write(row, col+4, d[2]);
        worksheet.write(row, col+5, d[3]);
        worksheet.write(row, col+6, d[4]);
        worksheet.write(row, col+7, d[5]);
        worksheet.write(row, col+8, d[6]);
        worksheet.write(row, col+9, d[7]);
        worksheet.write(row, col+10, "");
        worksheet.write(row, col+11, "");
        worksheet.write(row, col+12, "");
        worksheet.write(row, col+13, "");
        row += 1
    workbook.close()

def fetch_data(url):
    proxy = {'https': 'https://11115:7My2Ng@world.nohodo.com:6811'}
    try:
        res = sess.get(url, proxies=proxy)
    except Exception as e:
        print("type error: " + str(e))
    return res

def get_PageNo(res):
    soup = BeautifulSoup(response.text, 'lxml')
    try:
        PageDiv = soup.find("span", {'class': 'EstiloFechaCl'}).text
        regx = re.compile(r'\d+')
        no = regx.search(PageDiv)
        if no:
            products = int(no.group())
            Pages = int(products/20)
            if products % 20 > 0:
                Pages += 1
        else:
            Pages = 0
    except Exception as e:
        Pages = 0
    return Pages

def Extract_data(res, url):
    soup = BeautifulSoup(res.text, 'lxml')
    try:
        container = soup.find('div', {'id': 'dasddd'})
        block = container.find_all('div', {'class': 'cont_prod_fam'})
        for li in block:
            Name = li.find('a', {'class': 'EstCab_prod_port'}).text
            namepart = Name.split(" ")
            Manufacturer = namepart[0]
            ProdURL = li.find('a', {'class': 'EstCab_prod_port'})['href']
            try:
                promo = price = li.find('span', {'class': 'EstPrice_prod'}).text.replace('\n', '').replace('\t', '').replace('€', '')
            except Exception as e:
                promo = price = "Check for Price"
            Itemnumber = li.find('span', {'class': 'EstPrice_prod_port'}).text.replace("PN: ", "");
            mpn = Itemnumber
            DataOut.append([url, Itemnumber, mpn, Manufacturer, Name, ProdURL, price, promo])
    except Exception as e:
        DataOut.append(['', '', '', '', '', '', '', ''])


print("start")
LOB = ''
site = ''
country = ''
excel_To_List()
for url in input:
    print(url)
    indx = url.index('^')
    if indx != 0:
        LOB = url[:indx]
        url = url[indx+1:]
        splitURL = url.split('/')
        country = splitURL[2].replace("www.pcexpansion.", "").upper()
        site = splitURL[2].replace('www.', '').replace('.', '_')
    response = fetch_data(url)
    Pages = int(get_PageNo(response))
    count = 0
    for i in range(0, Pages):
        len = url.index('?rgi=')
        caturl = url[:len+5] + str(count)
        CatRes = fetch_data(caturl)
        Extract_data(CatRes, url)
        count += 20
Result_Out_Excel(DataOut)
