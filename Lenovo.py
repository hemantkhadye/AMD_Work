import requests
import os
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import xlsxwriter

InputURL = []

def Excel_To_List():
    InBook = load_workbook("E:\hemant\Hemant Python\AMD\Input.xlsx")
    InSheet = InBook.get_active_sheet()
    for row in InSheet.iter_rows():
        InputURL.append()

    # InputURL = [row.value for row in InSheet.rows]


def List_To_Excel():
    print('Done')

print("Started")
Excel_To_List()
url = 'https://www.amazon.co.uk/s/ref=sr_pg_1?rh=n%3A340831031%2Cn%3A428651031%2Ck%3Agaming%2Cp_6%3AA3P5ROKL5A1OLE&page=1&bbn=428651031&keywords=gaming&ie=UTF8&qid=1377150458'

proxy_support = {'https': 'https://69.65.48.225:80'}

CatHome = requests.get(url, proxies=proxy_support)